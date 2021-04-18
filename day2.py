# FUNCTIONS
# a = 111
# def my_func():
# 	a = 123
# 	print(a)
# 	print("hello world")
# 	return a/2
# # my_func()
# # print(a)
# print(my_func())

# def new_func(a,b):
# 	c = a + b
# 	print(c)
# new_func(1,2)

# def one_func(name,surname):
# 	a = (f"Hello {name} {surname}")
# 	print(a)
# one_func("Aya", "Aip")

# def one_func(x,y):
# 	print(x*y)
# one_func(3, 4)


# def my_func (a, b=2, c=3):
# 	print(a,b,c)
# a = 123
# my_func(a, c=9)
# PASS - Propusti 


# a = input("Write your age\n")
# try:
# 	a = int(a)
# 	age(a)
# except Exception as e: 
# 	print("Please write your age")
# def age(a):
# 	if a >17:
# 		print("Come in")
# 	else:
# 		print("Go out")

# import new_module as nmd
# print(nmd.a)
# nmd.age(9,7)

# import arithmetic as ar
# import str_work as sw 
# from arithmetic import (plus,minus)
# print(ar.plus(1,3))
# print(ar.minus(4,3))

# print(sw.add_str('asd', 'www'))

# # DATE TIME CALENDAR
# x = datetime.datetime(2020, 5, 17)
# print(x)
# print(x.strftime("%d.%m.%Y %H:%M"))

# from openpyxl import load_workbook
# wb = load_workbook('asd.xlsx')

# print(wb.sheetnames)

# sheet = wb.get_sheet_by_name('users')
# print(sheet)
# sheet['A2'].value='QQQQ'
# print(sheet['A2'].column)
# print(sheet['A2'].row)
# print(sheet['A2'].coordinate)

# for i in range(1,4):
# 	print(i,sheet.cell(row=i, column=1).value)
# wb.save('zzzz.xlsx')

from openpyxl import load_workbook
wb = load_workbook('salary.xlsm')

print(wb.sheetnames)

sheet = wb.get_sheet_by_name('Database')
sum1 = 0
sum1 = int(sum1)
for i in range(4,25):
	a = sheet.cell(row=i, column=11).value
	sum1 = sum1 + int(a)

print(sum1)
	# # wb.save('Salar.xlsx')
